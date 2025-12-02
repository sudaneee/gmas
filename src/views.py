from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib import messages
from django.contrib.auth import authenticate
from django.contrib.auth.models import User, auth
from django.contrib.auth.decorators import login_required
from django.contrib.humanize.templatetags import humanize
from django.template.loader import render_to_string
from src.models import (
    Subject,
    StudentClass,
    Session,
    Term,
    Student,
    StudentResult,
    StudentBehaviouralAssessment,
    signature,
    sets,
)
from django.db import connection
import datetime
from collections import defaultdict
import openpyxl
from io import BytesIO
from openpyxl import Workbook
from natural import number
# from weasyprint import HTML

from collections import defaultdict
import os
from django.conf import settings
from django.template.loader import render_to_string

from xhtml2pdf import pisa

import os
from collections import defaultdict
import datetime





# cursor to move around the database

c = connection.cursor()



def home(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']

        user = auth.authenticate(username=username, password=password)
        if user is not None:
            auth.login(request, user)
            return render(request, 'src/dashboard.html')

        else:
            messages.error(request, 'Incorrect Login-in Details')
            return render(request, 'src/welcome.html')

    else:
        return render(request, 'src/welcome.html')
def side_nav(request):
    return render(request, 'src/index.html')
def ResultUpload(request):
    if request.user.is_authenticated:
        classes = StudentClass.objects.all()
        subjects = Subject.objects.all()
        context = {
            'subjects': subjects,
            'classes': classes
        }
        if "GET" == request.method:
            return render(request, 'src/stdupload.html', context)
        else:

            session = request.POST['session']
            term = request.POST['term']
            classs = request.POST['class']
            subject = request.POST['subject']

            #---------------Getting IDs of the Subjects, Session, and Class---------------------
            ses = Session.objects.get(session_name=session)
            session_id = ses.id

            trm = Term.objects.get(term_name=term)
            term_id = trm.id

            clas = StudentClass.objects.get(class_name=classs)
            class_id = clas.id

            subj = Subject.objects.get(subject_name=subject)
            subject_id = subj.id

            excel_file = request.FILES["excel_file"]

            # you may put validations here to check extension or file size

            wb = openpyxl.load_workbook(excel_file)

            # getting a particular sheet by name out of many sheets
            worksheet = wb["result"]
            ids = []
            ca1s = []
            ca2s = []
            exams = []
            ses_ids = []
            trm_ids = []
            clas_ids = []
            subj_ids =[]
            total_score = []
            for i, j, k, l in zip((worksheet['A']), (worksheet['C']), (worksheet['D']), (worksheet['E'])):
                ids.append(i.value)
                ca1s.append(j.value)
                ca2s.append(k.value)
                exams.append(l.value)
                ses_ids.append(session_id)
                trm_ids.append(term_id)
                clas_ids.append(class_id)
                subj_ids.append(subject_id)


            #------------removing first row of the excel-----------
            f_ids = ids[1:]
            f_ca1s = ca1s[1:]
            f_ca2s = ca2s[1:]
            f_exams = exams[1:]

            #-----------coverting to interger list----------------
            a_ids = [int(i) for i in f_ids]
            a_ca1s = [int(i) for i in f_ca1s]
            a_ca2s = [int(i) for i in f_ca2s]
            a_exams = [int(i) for i in f_exams]

            #--------------sumation of the CA to get Total Marks
            for i in range(len(f_ca1s)):
                total_score.append(a_ca1s[i] + a_ca2s[i] + a_exams[i])
            print(total_score)

            position = []
            for total in total_score:
                count = 0
                for pos in total_score:
                    if pos > total:
                        count += 1
                position.append(count +1)


            clso = int(class_id)
            sess = int(session_id)
            trm = int(term_id)
            sbj = int(subject_id)

            c = connection.cursor()



            queryset = StudentResult.objects.filter(session=sess, term=trm, student_class=clso, subject=sbj).exists()
            if queryset:
                print("record is already there")
            else:

                c.executemany('INSERT INTO src_studentresult (ca1, ca2, exams, total, subject_position, student_id, student_class_id, subject_id, session_id, term_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)', zip(a_ca1s, a_ca2s, a_exams, total_score, position, a_ids, clas_ids, subj_ids, ses_ids, trm_ids))
                connection.commit()
                c.close()
                messages.info(request, 'Result Uploaded Successfully')

            s = "ZMS/SS/12"



        return render(request,'src/stdupload.html', context)
    else:
        return render(request, 'src/welcome.html')


def stdcreate(request):
    if request.user.is_authenticated:
        classes = StudentClass.objects.all()
        context = {
            'classes': classes
        }



        if "GET" == request.method:
            return render(request, 'src/stdcreate.html', context)
        else:
            classs = request.POST['class']


            #---------------Getting IDs of the Class---------------------

            clas = StudentClass.objects.get(class_name=classs)
            class_id = clas.id

            excel_file = request.FILES["excel_file"]

            # you may put validations here to check extension or file size

            wb = openpyxl.load_workbook(excel_file)

            # getting a particular sheet by name out of many sheets
            worksheet = wb["Sheet1"]
            print(worksheet)
            names = []
            genders = []


            for i, j in zip(worksheet['A'], worksheet['B']):
                names.append(i.value)
                genders.append(j.value)
            #------------removing first row of the excel-----------

            f_names = names[0:]
            f_genders = genders[0:]




            for i,j in zip(f_names, f_genders):
                Student.objects.create(
                    student_name = i,
                    gender = j,
                    student_class = StudentClass.objects.get(class_name=classs)
                )


            messages.info(request, 'Record Uploaded Successfully')




        return render(request, 'src/stdcreate.html', context)
    else:
        return render(request, 'src/welcome.html')

def resultcreate(request):
    global students
    global subject
    global term
    global session
    global classs
    global class_id
    global session_id
    global term_id
    global subject_id
    global student_id

    if request.user.is_authenticated:

        if request.method == 'POST' and "form1" in request.POST:

            session = request.POST['session']
            term = request.POST['term']
            classs = request.POST['class']
            subject = request.POST['subject']

            ses = Session.objects.get(session_name=session)
            session_id = ses.id

            trm = Term.objects.get(term_name=term)
            term_id = trm.id

            clas = StudentClass.objects.get(class_name=classs)
            class_id = clas.id

            subj = Subject.objects.get(subject_name=subject)
            subject_id = subj.id

            students = Student.objects.filter(student_class=class_id, status=True).all()
            student_id = []
            for i in students:
                student_id.append(i.id)


            subjects = Subject.objects.all()
            classes = StudentClass.objects.all()
            sessions = Session.objects.all()
            terms = Term.objects.all()
            context = {
                'subjects': subjects,
                'classes': classes,
                'students': students
            }

            return render(request, 'src/resultcreate.html', context)

        if request.method == 'POST' and "form2" in request.POST:
            ses_ids = []
            trm_ids = []
            clas_ids = []
            subj_ids =[]
            total_score = []

            #ids = request.POST.getlist('id')
            names = request.POST.getlist('name')
            ca1s = request.POST.getlist('ca1')
            ca1s2 = [int(i) for i in ca1s]
            ca2s = request.POST.getlist('ca2')
            ca2s2 = [int(i) for i in ca2s]
            exams = request.POST.getlist('exams')
            exams2 = [int(i) for i in exams]
            for i in range(len(ca1s2)):
                total_score.append(ca1s2[i] + ca2s2[i] + exams2[i])

            position = []
            for total in total_score:
                count = 0
                for pos in total_score:
                    if pos > total:
                        count += 1
                position.append(count +1)

            for i in students:
                ses_ids.append(session_id)
                trm_ids.append(term_id)
                clas_ids.append(class_id)
                subj_ids.append(subject_id)

            clso = int(class_id)
            sess = int(session_id)
            trm = int(term_id)
            sbj = int(subject_id)

            c = connection.cursor()

            queryset = StudentResult.objects.filter(session=sess, term=trm, student_class=clso, subject=sbj).exists()
            if queryset:
                print("record is already there")
            else:
                c.executemany('INSERT INTO src_studentresult (ca1, ca2, exams, total, subject_position, student_id, student_class_id, subject_id, session_id, term_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)', zip(ca1s, ca2s, exams, total_score, position, student_id, clas_ids, subj_ids, ses_ids, trm_ids))
                connection.commit()
                c.close()
                messages.info(request, 'Result Uploaded Successfully')




            subjects = Subject.objects.all()
            classes = StudentClass.objects.all()
            sessions = Session.objects.all()
            terms = Term.objects.all()
            context = {
                'subjects': subjects,
                'classes': classes,
                'students': students
            }

            return render(request, 'src/resultcreate.html', context)


        else:
            subjects = Subject.objects.all()
            classes = StudentClass.objects.all()
            context = {
                'subjects': subjects,
                'classes': classes
            }
            return render(request, 'src/resultcreate.html', context)
    else:
        return render(request, 'src/welcome.html')

def updateresult(request):
    global students
    global subject
    global term
    global session
    global classs
    global class_id
    global session_id
    global term_id
    global subject_id
    global result_ids
    if request.user.is_authenticated:

        if request.method == 'POST' and "form1" in request.POST:

            session = request.POST['session']
            term = request.POST['term']
            classs = request.POST['class']
            subject = request.POST['subject']

            ses = Session.objects.get(session_name=session)
            session_id = ses.id

            trm = Term.objects.get(term_name=term)
            term_id = trm.id

            clas = StudentClass.objects.get(class_name=classs)
            class_id = clas.id

            subj = Subject.objects.get(subject_name=subject)
            subject_id = subj.id

            students = Student.objects.filter(student_class=class_id).all()
            student_id = []




            all_results = StudentResult.objects.filter(session=session_id, term=term_id, student_class=class_id,subject=subject_id).all()

            result_ids = StudentResult.objects.filter(session=session_id, term=term_id, student_class=class_id,subject=subject_id).values_list('id', flat=True)
            print(result_ids)


            for i in students:
                student_id.append(i.id)

            subjects = Subject.objects.all()
            classes = StudentClass.objects.all()
            sessions = Session.objects.all()
            terms = Term.objects.all()

            result_students_list = zip(all_results, students)
            context = {
                'subjects': subjects,
                'classes': classes,
                'all_results': all_results,
                'result_students_list': result_students_list
            }

            return render(request, 'src/updateresult.html', context)

        if request.method == 'POST' and "form2" in request.POST:
            ses_ids = []
            trm_ids = []
            clas_ids = []
            subj_ids =[]
            total_score = []

            ids = request.POST.getlist('id')
            names = request.POST.getlist('name')
            ca1s = request.POST.getlist('ca1')
            ca1s2 = [int(i) for i in ca1s]
            ca2s = request.POST.getlist('ca2')
            ca2s2 = [int(i) for i in ca2s]
            exams = request.POST.getlist('exams')
            exams2 = [int(i) for i in exams]
            for i in range(len(ca1s2)):
                total_score.append(ca1s2[i] + ca2s2[i] + exams2[i])
            print(total_score)

            position = []
            for total in total_score:
                count = 0
                for pos in total_score:
                    if pos > total:
                        count += 1
                position.append(count +1)


            print(ca1s)

            for i in students:
                ses_ids.append(session_id)
                trm_ids.append(term_id)
                clas_ids.append(class_id)
                subj_ids.append(subject_id)

            c = connection.cursor()

            c.executemany("""UPDATE src_studentresult SET ca1=?, ca2=?, exams=?, total=?, subject_position=? WHERE
                id=? AND student_class_id=? AND subject_id=? AND session_id=? AND term_id=?""",
                zip(ca1s, ca2s, exams, total_score, position, result_ids, clas_ids, subj_ids, ses_ids, trm_ids))
            connection.commit()
            c.close()
            messages.info(request, 'Result Updated Successfully')

            subjects = Subject.objects.all()
            classes = StudentClass.objects.all()
            sessions = Session.objects.all()
            terms = Term.objects.all()
            context = {
                'subjects': subjects,
                'classes': classes,
                'students': students
            }

            return render(request, 'src/updateresult.html', context)
        else:
            subjects = Subject.objects.all()
            classes = StudentClass.objects.all()
            context = {
                'subjects': subjects,
                'classes': classes
            }
            return render(request, 'src/updateresult.html', context)
    else:
        return render(request, 'src/welcome.html')


def single_result_update(request):
    global students
    global subject
    global term
    global session
    global classs
    global class_id
    global session_id
    global term_id
    global subject_id
    global student_qr_id
    global student_id


    if request.user.is_authenticated:

        if request.method == 'POST' and "form1" in request.POST:
            session = request.POST['session']
            term = request.POST['term']
            classs = request.POST['class']
            subject = request.POST['subject']
            student_id = request.POST['student_id']

            ses = Session.objects.get(session_name=session)
            session_id = ses.id

            trm = Term.objects.get(term_name=term)
            term_id = trm.id

            clas = StudentClass.objects.get(class_name=classs)
            class_id = clas.id

            subj = Subject.objects.get(subject_name=subject)
            subject_id = subj.id

            student = []
            student.append(student_id)


            get_student = StudentResult.objects.get(student=student_id, session=session_id, term=term_id, student_class=class_id, subject=subject_id)
            subjects = Subject.objects.all()
            classes = StudentClass.objects.all()
            context = {
                'subjects': subjects,
                'classes': classes,
                'get_student': get_student,
                'student_id': student_id
                }
            return render(request, 'src/single_result_update.html', context)




        if request.method == 'POST' and "form2" in request.POST:


            ca1 = request.POST['ca1']
            ca2 = request.POST['ca2']
            exams = request.POST['exams']

            total_score = int(ca1) + int(ca2) + int(exams)



            queryy = StudentResult.objects.get(student=student_id, session=session_id, term=term_id, student_class=class_id, subject=subject_id)

            queryy.ca1 = ca1
            queryy.save()
            queryy.ca2 = ca2
            queryy.save()
            queryy.exams = exams
            queryy.save()
            queryy.total = total_score
            queryy.save()


            subjects = Subject.objects.all()
            classes = StudentClass.objects.all()
            context = {
                'subjects': subjects,
                'classes': classes
                }
            return render(request, 'src/single_result_update.html', context)


        else:
            subjects = Subject.objects.all()
            classes = StudentClass.objects.all()
            context = {
                'subjects': subjects,
                'classes': classes
                }
            return render(request, 'src/single_result_update.html', context)
    else:
        return render(request, 'src/welcome.html')



# Example:
# from .models import StudentClass, Session, Term, StudentResult, StudentBehaviouralAssessment, signature, sets

# # @login_required
# def result_view(request):
#     # ReportLab constants
#     full_width_inches = 7.5 * inch
#     # Vertical space calculation
#     A4_HEIGHT = 841.8976  # A4 height in points
#     TOP_MARGIN = 30
#     BOTTOM_MARGIN = 20
#     # The total vertical space available for content (in points)
#     CONTENT_HEIGHT_PT = A4_HEIGHT - TOP_MARGIN - BOTTOM_MARGIN
    
#     # --- Initial Page/Error Handling ---
#     if request.method != 'POST' or "form1" not in request.POST:
#         classes = StudentClass.objects.all()
#         return render(request, 'src/result_view.html', {'classes': classes})

#     try:
#         session_name = request.POST['session']
#         term_name = request.POST['term']
#         class_name = request.POST['class']
#         ses = Session.objects.get(session_name=session_name)
#         trm = Term.objects.get(term_name=term_name)
#         clas = StudentClass.objects.get(class_name=class_name)
#     except Exception as e:
#         classes = StudentClass.objects.all()
#         return render(request, 'src/result_view.html', {
#             'classes': classes,
#             'error': f'Database lookup error: {e}'
#         })

#     # --- Data Retrieval ---
#     all_results = StudentResult.objects.filter(
#         session=ses, term=trm, student_class=clas
#     ).order_by('student__id')

#     if not all_results.exists():
#         classes = StudentClass.objects.all()
#         return render(request, 'src/result_view.html', {
#             'classes': classes,
#             'error': 'No results found for selection.'
#         })

#     students_data = defaultdict(lambda: {'subjects': [], 'total': 0, 'num_subjects': 0})
#     for result in all_results:
#         student_id = result.student.id
#         data = students_data[student_id]
#         if 'student' not in data:
#             data['student'] = result.student
#             data['classs'] = result.student_class.class_name
#         data['subjects'].append({
#             'subject': result.subject.subject_name,
#             'ca1': result.ca1,
#             'ca2': result.ca2,
#             'exams': result.exams,
#             'total': result.total,
#             'subject_position': result.subject_position,
#         })
#         data['total'] += result.total
#         data['num_subjects'] += 1

#     allAvr = []
#     for data in students_data.values():
#         average = data['total'] / data['num_subjects'] if data['num_subjects'] > 0 else 0
#         data['average'] = round(average, 2)
#         data['overal_total'] = data['total']
#         allAvr.append(data['average'])

#     sorted_unique_averages = sorted(list(set(allAvr)), reverse=True)
#     position_map = {avg: humanize.ordinal(idx + 1) for idx, avg in enumerate(sorted_unique_averages)}

#     for data in students_data.values():
#         data['c_pos'] = position_map.get(data['average'], '-')

#     f_classs = str(clas.class_name).upper()
#     date = datetime.datetime.now()
#     no_in_class = len(students_data)

#     pre_primary_classes = ["NURSERY", "KINDERGARTEN", "KG", "BASIC", "PN"]
#     is_pre_primary = any(keyword in f_classs for keyword in pre_primary_classes)

#     # Get signature based on class type
#     if is_pre_primary:
#         signs_obj = signature.objects.filter(classs=clas).first()
#     else:
#         signs_obj = signature.objects.filter(classs=clas).first()

#     bhv_results = StudentBehaviouralAssessment.objects.filter(session=ses, term=trm, student_class=clas)
#     bhv_dict = {b.student_id: b for b in bhv_results}

#     final_results_list = []
#     for student_id, result in students_data.items():
#         result['bhv'] = bhv_dict.get(student_id)
#         final_results_list.append(result)

#     some_images = sets.objects.first()

#     # --- ReportLab PDF Setup ---
#     response = HttpResponse(content_type='application/pdf')
#     filename = f'Examination_Report_{f_classs.replace(" ", "_")}_{date.strftime("%Y%m%d")}.pdf'
#     response['Content-Disposition'] = f'inline; filename="{filename}"'

#     doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=20, leftMargin=20, topMargin=30, bottomMargin=20)
#     elements = []

#     # --- Define Custom Styles Correctly ---
#     styles = getSampleStyleSheet()

#     normal_wrap_style = styles['Normal'].clone('NormalWrap')
#     normal_wrap_style.fontSize = 9
#     normal_wrap_style.leading = 11
#     styles.add(normal_wrap_style)

#     center_wrap_style = normal_wrap_style.clone('CenterWrap')
#     center_wrap_style.alignment = 1 
#     styles.add(center_wrap_style)

#     right_wrap_style = normal_wrap_style.clone('RightWrap')
#     right_wrap_style.alignment = 2
#     styles.add(right_wrap_style)

#     heading3_center = styles['Heading3'].clone('Heading3Center')
#     heading3_center.alignment = 1
#     styles.add(heading3_center)

#     title_center = styles['Title'].clone('TitleCenter')
#     title_center.alignment = 1
#     title_center.fontSize = 16
#     styles.add(title_center)

#     # --- Common Table Style ---
#     common_style = [ 
#         ('GRID', (0, 0), (-1, -1), 1, colors.black),
#         ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
#     ]
    
#     # --- Estimate Fixed Height Elements (in points) ---
#     # We now estimate only the static heights that DON'T need flexibility:
    
#     # Header Image + Title Block (0.8in + title height + 0.1in total approx)
#     H_HEADER_BLOCK = (0.8 * inch) + (0.3 * inch) 
    
#     # Student Info Table (approx 4 rows x 15pt/row)
#     H_STUDENT_INFO = (4 * 15) 
    
#     # Comment Table (approx 1 row x 15pt)
#     H_COMMENT_TABLE = (1 * 15) 
    
#     # Signature Table (approx 2 rows x 15pt + image height 0.4in)
#     H_SIGNATURE_TABLE = (2 * 15) + (0.4 * inch)
    
#     # Static Spacing (removed most, keeping a small buffer)
#     H_STATIC_SPACING = (0.2 * inch) 
    
#     H_FIXED_ELEMENTS_MIN = H_HEADER_BLOCK + H_STUDENT_INFO + H_COMMENT_TABLE + H_SIGNATURE_TABLE + H_STATIC_SPACING 
    
#     # --- Loop through each student result ---
#     for result in final_results_list:
        
#         # --- DYNAMIC HEIGHT CALCULATION ---
        
#         # 1. Check if Behavioural Assessment is present and add its minimum height
#         bhv_is_present = bool(result['bhv'])
#         H_BHV_TABLE_MIN = (4 * 15) if bhv_is_present else 0
        
#         # 2. Total minimum height for non-scores elements
#         H_TOTAL_NON_SCORES_MIN = H_FIXED_ELEMENTS_MIN + H_BHV_TABLE_MIN

#         # 3. Calculate available space for the Scores Table
#         available_space_pt = CONTENT_HEIGHT_PT - H_TOTAL_NON_SCORES_MIN
        
#         # 4. Calculate required row height
#         num_scores_rows = len(result['subjects']) + 1 # Subjects + Footer row
#         min_row_height = available_space_pt / num_scores_rows
#         # Set a reasonable minimum to prevent rows from becoming too small
#         min_row_height = max(min_row_height, 20) 
        
#         # --- Header ---
#         if some_images and some_images.h_image and some_images.h_image.name:
#             image_path = os.path.join(settings.MEDIA_ROOT, some_images.h_image.name)
#             if os.path.exists(image_path):
#                 img = Image(image_path)
#                 img._width = full_width_inches
#                 img._height = 0.8 * inch
#                 img.hAlign = 'CENTER'
#                 elements.append(img)

#         elements.append(Paragraph("<b>EXAMINATION REPORT SHEET</b>", styles['TitleCenter']))
        
#         # --- Student Info Table ---
#         student_data = [
#             [Paragraph("<b>Student Information</b>", styles['Heading3Center']), '', '', ''],
#             ["Name:", Paragraph(str(result['student']), normal_wrap_style), "Class:", Paragraph(result['classs'], normal_wrap_style)],
#             ["Session:", Paragraph(str(ses), normal_wrap_style), "Term:", Paragraph(str(trm), normal_wrap_style)],
#             ["No. in Class:", Paragraph(str(no_in_class), normal_wrap_style), "Gender:", Paragraph(str(result['student'].gender), normal_wrap_style)],
#         ]

#         if not is_pre_primary:
#             student_data[3].insert(2, "Class Position:")
#             student_data[3].insert(3, Paragraph(str(result['c_pos']), normal_wrap_style)) 
#             col_widths_student = [1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch] 
#             student_data[0] = [student_data[0][0], '', '', '', '']
#             student_table_style = TableStyle(common_style + [
#                 ('SPAN', (0,0), (-1,0)),
#                 ('ALIGN', (1,1), (-1,-1), 'CENTER'),
#                 ('BACKGROUND', (0,1), (0,-1), colors.lightgrey),
#                 ('BACKGROUND', (2,1), (2,-1), colors.lightgrey),
#                 ('BACKGROUND', (4,1), (4,-1), colors.lightgrey),
#             ])
#         else:
#             col_widths_student = [1.5*inch, 2.25*inch, 1.5*inch, 2.25*inch] 
#             student_table_style = TableStyle(common_style + [
#                 ('SPAN', (0,0), (-1,0)),
#                 ('ALIGN', (1,1), (-1,-1), 'CENTER'),
#                 ('BACKGROUND', (0,1), (0,-1), colors.lightgrey),
#                 ('BACKGROUND', (2,1), (2,-1), colors.lightgrey),
#             ])

#         student_table = Table(student_data, colWidths=col_widths_student)
#         student_table.setStyle(student_table_style)
#         elements.append(student_table)

#         # --- Scores Table (DYNAMIC HEIGHT) ---
#         scores_header = [
#             Paragraph("S/N", center_wrap_style), 
#             Paragraph("<b>Subject</b>", center_wrap_style), 
#             Paragraph("<b>1st CA</b>", center_wrap_style), 
#             Paragraph("<b>2nd CA</b>", center_wrap_style), 
#             Paragraph("<b>Exam</b>", center_wrap_style), 
#             Paragraph("<b>Total</b>", center_wrap_style), 
#             Paragraph("<b>Grade</b>", center_wrap_style)
#         ]
        
#         if not is_pre_primary:
#             scores_header.append(Paragraph("<b>Position</b>", center_wrap_style))
#             col_widths_scores = [0.5*inch, 2.3*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.9*inch, 1.0*inch]
#         else:
#             col_widths_scores = [0.5*inch, 2.5*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 1.7*inch] 

#         scores_data = [scores_header]

#         for idx, subject in enumerate(result['subjects']):
#             total = subject['total']
#             if total <= 40: grade = 'E'
#             elif 41 <= total <= 49: grade = 'D'
#             elif 50 <= total <= 59: grade = 'C'
#             elif 60 <= total <= 69: grade = 'B'
#             elif 70 <= total <= 80: grade = 'A'
#             elif 81 <= total <= 100: grade = 'A+'
#             else: grade = '-'

#             row = [
#                 Paragraph(str(idx + 1), center_wrap_style), 
#                 Paragraph(subject['subject'], normal_wrap_style), 
#                 Paragraph(str(subject['ca1']), center_wrap_style), 
#                 Paragraph(str(subject['ca2']), center_wrap_style), 
#                 Paragraph(str(subject['exams']), center_wrap_style), 
#                 Paragraph(str(total), center_wrap_style), 
#                 Paragraph(grade, center_wrap_style)
#             ]
#             if not is_pre_primary:
#                 row.append(Paragraph(str(subject['subject_position']), center_wrap_style))
#             scores_data.append(row)

#         # Footer
#         footer_row = [''] * len(scores_header)
#         footer_row[2] = Paragraph('Overall Total:', right_wrap_style)
#         footer_row[3] = Paragraph(str(result['overal_total']), center_wrap_style)
#         footer_row[4] = Paragraph('Average:', right_wrap_style)
#         footer_row[5] = Paragraph(str(result['average']), center_wrap_style)
#         scores_data.append(footer_row)

#         scores_table_style = TableStyle(common_style + [
#             ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
#             ('ALIGN', (0,1), (-1,-1), 'CENTER'),
#             ('VALIGN', (0,-1), (-1,-1), 'MIDDLE'),
#             ('SPAN', (0,-1), (1,-1)), 
#             ('MINROWHEIGHTS', (0, 1), (-1, -2), min_row_height), 
#             ('MINROWHEIGHTS', (0, 0), (-1, 0), 20), 
#             ('MINROWHEIGHTS', (0, -1), (-1, -1), 20), 
#         ])
        
#         scores_table = Table(scores_data, colWidths=col_widths_scores)
#         scores_table.setStyle(scores_table_style)
#         elements.append(scores_table)

#         # --- Behavioural Assessment ---
#         bhv = result['bhv']
#         if bhv:
#             bhv_data = [
#                 [Paragraph("<b>Behavioural Assessment</b>", styles['Heading3Center']), '', '', '', '', '', '', ''],
#                 ["Conduct", Paragraph(str(bhv.conduct), normal_wrap_style), "Punctuality", Paragraph(str(bhv.punctuality), normal_wrap_style), "Dedication", Paragraph(str(bhv.dedication), normal_wrap_style), "Participation", Paragraph(str(bhv.participation), normal_wrap_style)],
#                 ["Hospitality", Paragraph(str(bhv.hospitality), normal_wrap_style), "Neatness", Paragraph(str(bhv.neatness), normal_wrap_style), "Creativity", Paragraph(str(bhv.creativity), normal_wrap_style), "Physical Health", Paragraph(str(bhv.physical), normal_wrap_style)],
#                 ["Days Opened", Paragraph(str(bhv.school_opened), normal_wrap_style), "Days Present", Paragraph(str(bhv.days_present), normal_wrap_style), "Days Absent", Paragraph(str(bhv.days_absent), normal_wrap_style), "Resumption", Paragraph(str(bhv.next_date_of_resumption), normal_wrap_style)],
#             ]
#             bhv_col_widths = [0.9*inch, 0.975*inch] * 4 
#             bhv_table = Table(bhv_data, colWidths=bhv_col_widths)
#             bhv_table.setStyle(TableStyle(common_style + [
#                 ('SPAN', (0,0), (-1,0)),
#                 ('ALIGN', (1,1), (-1,-1), 'CENTER'), 
#                 ('BACKGROUND', (0,1), (0,-1), colors.lightgrey),
#                 ('BACKGROUND', (2,1), (2,-1), colors.lightgrey),
#                 ('BACKGROUND', (4,1), (4,-1), colors.lightgrey),
#                 ('BACKGROUND', (6,1), (6,-1), colors.lightgrey),
#             ]))
#             elements.append(bhv_table)
        
#         # --- GAP FILLER & FINAL BLOCK ---
#         # The key to pushing signatures to the bottom is using a table as a vertical spacer.
        
#         # Calculate the actual height remaining for the gap filler.
#         # This is complex and relies on accurate estimation. A simpler, more robust way
#         # in ReportLab is often to use the Frame/PageTemplate concept, but for simplicity
#         # here, we will wrap the final elements in a table that attempts to fill the remainder.

#         # Calculate the cumulative height of all elements added so far (approximate)
#         # This approach is generally unreliable in ReportLab without custom flowables.
#         # A simpler, more reliable method for "bottom of page" placement is to use a large table,
#         # but since we already used all available space with the dynamic scores table, 
#         # we now need to ensure the final elements are not placed too far down if the scores table is small.
        
#         # Re-introducing a calculated space, based on the scores table min height.
#         if num_scores_rows < 10: # If there are few subjects, add vertical padding
            
#             # The height used by the Scores Table is num_scores_rows * min_row_height
#             H_SCORES_TABLE_USED = num_scores_rows * min_row_height
            
#             # Calculate the total height taken by all elements so far (including scores table)
#             H_TOTAL_USED_MIN = H_TOTAL_NON_SCORES_MIN + H_SCORES_TABLE_USED
            
#             # The space left until the page bottom
#             H_GAP_SPACE = CONTENT_HEIGHT_PT - H_TOTAL_USED_MIN - H_STATIC_SPACING # Subtract final spacing for comment/sig

#             # Use a dummy table to fill this gap
#             if H_GAP_SPACE > 0:
#                 gap_table_data = [['']]
#                 gap_table = Table(gap_table_data, colWidths=[full_width_inches])
#                 gap_table.setStyle(TableStyle([
#                     ('MINROWHEIGHTS', (0, 0), (0, 0), H_GAP_SPACE),
#                     ('GRID', (0, 0), (0, 0), 0.25, colors.white), # Invisible grid
#                 ]))
#                 elements.append(gap_table)


#         # --- General Comment ---
#         avg = result['average']
#         if avg <= 49:
#             comment = "Well done, put more effort."
#         elif 50 <= avg <= 59:
#             comment = "Don't relent in your studies, you can do better."
#         elif 60 <= avg <= 69:
#             comment = "A very good performance."
#         elif 70 <= avg <= 80:
#             comment = "A brilliant performance, keep it up."
#         elif 81 <= avg <= 100:
#             comment = "Bravo! An outstanding and excellent performance, keep it up."
#         else:
#             comment = "Invalid score."

#         comment_data = [
#             [Paragraph("<b>General Comment:</b>", normal_wrap_style), Paragraph(comment, normal_wrap_style)]
#         ]
#         comment_table = Table(comment_data, colWidths=[1.5*inch, 6.0*inch]) 
#         comment_table.setStyle(TableStyle(common_style))
#         elements.append(comment_table)
#         elements.append(Spacer(1, 0.2 * inch)) # Retain small final spacer

#         # --- Signatures ---
#         if is_pre_primary:
#             sign_path = (
#                 os.path.join(settings.MEDIA_ROOT, signs_obj.t_image.name)
#                 if signs_obj and signs_obj.t_image and signs_obj.t_image.name and
#                 os.path.exists(os.path.join(settings.MEDIA_ROOT, signs_obj.t_image.name))
#                 else None
#             )
#             sign_label = "Class Teacher's Signature"
#         else:
#             sign_path = (
#                 os.path.join(settings.MEDIA_ROOT, signs_obj.p_image.name)
#                 if signs_obj and signs_obj.p_image and signs_obj.p_image.name and
#                 os.path.exists(os.path.join(settings.MEDIA_ROOT, signs_obj.p_image.name))
#                 else None
#             )
#             sign_label = "Principal's Signature"

#         if sign_path:
#             try:
#                 sign_img = Image(sign_path, 1.2*inch, 0.4*inch)
#             except:
#                 sign_img = Paragraph("________________", center_wrap_style)
#         else:
#             sign_img = Paragraph("________________", center_wrap_style)

#         signature_data = [
#             [sign_img, Paragraph(date.strftime("%d-%m-%Y"), center_wrap_style)],
#             [Paragraph(sign_label, center_wrap_style), Paragraph("Date", center_wrap_style)]
#         ]
#         signature_table = Table(signature_data, colWidths=[3.75*inch, 3.75*inch]) 
#         signature_table.setStyle(TableStyle([
#             ('ALIGN', (0,0), (-1,-1), 'CENTER'),
#             ('VALIGN', (0,0), (-1,-1), 'BOTTOM'),
#             ('BOTTOMPADDING', (0,0), (-1,-1), 0),
#         ]))
#         elements.append(signature_table)

#         if result != final_results_list[-1]:
#             elements.append(PageBreak())

#     # --- Build PDF ---
#     doc.build(elements)
#     return response



# You must ensure Session, Term, StudentClass, StudentResult, 
# StudentBehaviouralAssessment, signature, and sets are imported from your models.
# from your_app.models import ... 


# --- ASSUME THESE IMPORTS ARE CORRECTLY DEFINED IN YOUR views.py ---
# from your_app.models import Session, Term, StudentClass, StudentResult, StudentBehaviouralAssessment, signature, sets
# Make sure to import all required models at the top of your actual file!
# -------------------------------------------------------------------



# -----------------------------------------
# GRADING FUNCTION (UPDATED)
# -----------------------------------------
def get_grade(total, is_kgn_nur):
    """
    total = numeric score
    is_kgn_nur = True for Kindergarten/Nursery
    """

    if is_kgn_nur:
        # KINDERGARTEN / NURSERY GRADING (OLD SYSTEM)
        if total <= 40: return 'E'
        elif 41 <= total <= 49: return 'D'
        elif 50 <= total <= 59: return 'C'
        elif 60 <= total <= 69: return 'B'
        elif 70 <= total <= 80: return 'A'
        elif 81 <= total <= 100: return 'A+'
        return '-'

    else:
        # OTHER CLASSES (NEW SYSTEM WITH F)
        if total < 40: return 'F'
        elif 40 <= total <= 49: return 'E'
        elif 50 <= total <= 59: return 'D'
        elif 60 <= total <= 69: return 'C'
        elif 70 <= total <= 79: return 'B'
        elif 80 <= total <= 100: return 'A'
        return '-'


def get_comment(avg):
    if avg <= 49: return "Well done, put more effort."
    elif 50 <= avg <= 59: return "Don't relent in your studies, you can do better."
    elif 60 <= avg <= 69: return "A very good performance."
    elif 70 <= avg <= 80: return "A brilliant performance, keep it up."
    elif 81 <= avg <= 100: return "Bravo! An outstanding performance."
    return "Invalid score."


# -----------------------------------------
# MAIN VIEW FUNCTION
# -----------------------------------------
def result_view(request):
    classes = StudentClass.objects.all()

    if request.method != "POST":
        return render(request, "src/result_view.html", {"classes": classes})

    # Fetch form data
    session = Session.objects.get(session_name=request.POST["session"])
    term = Term.objects.get(term_name=request.POST["term"])
    clas = StudentClass.objects.get(class_name=request.POST["class"])

    # Retrieve student results
    all_results = StudentResult.objects.filter(
        session=session,
        term=term,
        student_class=clas
    ).order_by("student__id")

    if not all_results.exists():
        return render(request, "src/result_view.html", {
            "classes": classes,
            "error": "No results found."
        })

    # -----------------------------------------
    # CLASS CATEGORY DETECTION
    # -----------------------------------------
    classname_upper = clas.class_name.upper()

    IS_KGN_NUR = any(x in classname_upper for x in ["KG", "KINDERGARTEN", "NURSERY"])
    IS_BASIC = "BASIC" in classname_upper

    USE_HEADTEACHER = IS_KGN_NUR or IS_BASIC

    # -----------------------------------------
    # GROUP RESULTS BY STUDENT
    # -----------------------------------------
    students_data = defaultdict(lambda: {
        "subjects": [],
        "total": 0,
        "num": 0,
    })

    for res in all_results:
        st = students_data[res.student.id]

        st["student"] = res.student
        st["class"] = res.student_class.class_name

        # Subject details
        st["subjects"].append({
            "subject": res.subject.subject_name,
            "ca1": res.ca1,
            "ca2": res.ca2,
            "exams": res.exams,
            "total": res.total,
            "grade": get_grade(res.total, IS_KGN_NUR),
            "pos": res.subject_position,
        })

        # Aggregate totals
        st["total"] += res.total
        st["num"] += 1

    # -----------------------------------------
    # COMPUTE AVERAGES + CLASS POSITION
    # -----------------------------------------
    all_avgs = []

    for st in students_data.values():
        avg = st["total"] / st["num"] if st["num"] else 0
        st["average"] = round(avg, 2)

        # GRADING depends on class type
        st["grade"] = get_grade(st["average"], IS_KGN_NUR)
        st["comment"] = get_comment(st["average"])

        all_avgs.append(st["average"])

    # Rank based on unique averages
    sorted_unique = sorted(list(set(all_avgs)), reverse=True)
    pos_map = {avg: humanize.ordinal(i + 1) for i, avg in enumerate(sorted_unique)}

    # -----------------------------------------
    # BEHAVIOUR + FINAL STRUCTURE
    # -----------------------------------------
    bhv_qs = StudentBehaviouralAssessment.objects.filter(
        session=session,
        term=term,
        student_class=clas
    )
    bhv_map = {b.student_id: b for b in bhv_qs}

    final_list = []

    for sid, st in students_data.items():
        st["position"] = pos_map.get(st["average"], "-")
        st["bhv"] = bhv_map.get(sid)
        final_list.append(st)

    # -----------------------------------------
    # SIGNATURE + HEADER IMAGE
    # -----------------------------------------
    signs = signature.objects.all().first()
    images = sets.objects.first()

    return render(request, "src/print_results.html", {
        "results": final_list,
        "session": session,
        "term": term,
        "class_name": clas.class_name,
        "no_in_class": len(final_list),

        "header_image": images.h_image.url if images and images.h_image else None,
        "signs": signs,
        "date": datetime.datetime.now().strftime("%d-%m-%Y"),

        # Flags for template logic
        "is_kgn_nur": IS_KGN_NUR,
        "is_basic": IS_BASIC,
        "use_headteacher_sign": USE_HEADTEACHER,
    })



def setting_position(request):
    if request.user.is_authenticated:

        c = connection.cursor()

        c.execute("""UPDATE src_studentresult SET subject_position=( SELECT count(*) FROM src_studentresult AS i WHERE i.total > src_studentresult.total AND i.student_class_id = src_studentresult.student_class_id AND i.subject_id = src_studentresult.subject_id AND i.session_id = src_studentresult.session_id AND i.term_id = src_studentresult.term_id) + 1;""")
        connection.commit()
        c.close()
        messages.info(request, 'Position Set Successfully')
        return render(request, 'src/dashboard.html')
    else:
        pass

def settings_(request):
    global students
    global subject
    global term
    global session
    global classs
    global class_id
    global session_id
    global term_id
    global subject_id
    global student_id
    if request.user.is_authenticated:



        if request.method == 'POST' and "form1" in request.POST:

            session = request.POST['session']
            term = request.POST['term']
            classs = request.POST['class']


            ses = Session.objects.get(session_name=session)
            session_id = ses.id

            trm = Term.objects.get(term_name=term)
            term_id = trm.id

            clas = StudentClass.objects.get(class_name=classs)
            class_id = clas.id

            students = Student.objects.filter(student_class=class_id).all()
            student_id = []

            for i in students:
                student_id.append(i.id)
            print(student_id)

            classes = StudentClass.objects.all()

            context = {
                'classes': classes,
                'students': students
                }

            return render(request, 'src/settings.html', context)

        if request.method == 'POST' and "form2" in request.POST:

            ses_ids = []
            trm_ids = []
            clas_ids = []

            ids = request.POST.getlist('id')
            conduct = request.POST.getlist('conduct')
            punc = request.POST.getlist('punctuality')
            ded = request.POST.getlist('dedication')
            part = request.POST.getlist('participation')
            hosp = request.POST.getlist('hospitality')
            creat= request.POST.getlist('creativity')
            phy = request.POST.getlist('physical')
            neat = request.POST.getlist('neatness')

            for i in student_id:
                ses_ids.append(session_id)
                trm_ids.append(term_id)
                clas_ids.append(class_id)



            queryset = StudentBehaviouralAssessment.objects.filter(session=session_id, term=term_id, student_class=class_id).exists()
            if queryset:
                print("record is already there")
            else:
                c = connection.cursor()
                c.executemany("""INSERT INTO src_studentbehaviouralassessment (conduct, punctuality, dedication, participation, hospitality,
                    neatness, creativity, physical, session_id, student_id, student_class_id, term_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    zip(conduct, punc, ded, part, hosp, neat, creat, phy, ses_ids, student_id, clas_ids, trm_ids))
                connection.commit()
                c.close()


            classes = StudentClass.objects.all()

            context = {
                'classes': classes
                }

            return render(request, 'src/settings.html', context)



        else:
            classes = StudentClass.objects.all()

            context = {
                'classes': classes
                }

            return render(request, 'src/settings.html', context)
    else:
        return render(request, 'src/welcome.html')


def single_result_view(request):

    if request.method != "POST" or "form1" not in request.POST:
        return render(request, "src/result_checker.html")

    session_name = request.POST["session"]
    term_name = request.POST["term"]
    student_id = request.POST["student_id"]

    # Fetch objects
    session = Session.objects.get(session_name=session_name)
    term = Term.objects.get(term_name=term_name)
    student = Student.objects.get(id=student_id)
    clas = student.student_class

    # ---------------------------------------------------------
    # CLASS TYPE LOGIC (KG/Nursery/Basic vs others)
    # ---------------------------------------------------------
    classname_upper = clas.class_name.upper()

    IS_KGN_NUR = any(x in classname_upper for x in ["KG", "KINDERGARTEN", "NURSERY"])
    IS_BASIC = "BASIC" in classname_upper

    USE_HEADTEACHER = IS_KGN_NUR or IS_BASIC

    # ---------------------------------------------------------
    # FETCH STUDENT RESULTS
    # ---------------------------------------------------------
    results = StudentResult.objects.filter(
        session=session,
        term=term,
        student=student
    )

    if not results.exists():
        return render(request, "src/result_checker.html", {"error": "No result found."})

    # ---------------------------------------------------------
    # PROCESS THIS STUDENT’S SCORES
    # ---------------------------------------------------------
    subjects = []
    total_score = 0

    for res in results:
        subjects.append({
            "subject": res.subject.subject_name,
            "ca1": res.ca1,
            "ca2": res.ca2,
            "exams": res.exams,
            "total": res.total,
            "grade": get_grade(res.total, IS_KGN_NUR),
            "pos": res.subject_position,
        })
        total_score += res.total

    average_display = round(total_score / len(subjects), 2)
    overall_grade = get_grade(average_display, IS_KGN_NUR)
    comment = get_comment(average_display)

    # ---------------------------------------------------------
    # CLASS POSITION (exact average used for ranking)
    # ---------------------------------------------------------
    class_results = StudentResult.objects.filter(
        session=session,
        term=term,
        student_class=clas
    )

    # Gather totals for all students in the class
    student_totals = defaultdict(lambda: {"total": 0, "num": 0, "student": None})

    for res in class_results:
        st = student_totals[res.student.id]
        st["student"] = res.student
        st["total"] += res.total
        st["num"] += 1

    # Compute averages for ranking (UNROUNDED)
    averages = []
    for rec in student_totals.values():
        avg = rec["total"] / rec["num"] if rec["num"] else 0
        rec["average"] = avg
        averages.append(avg)

    # Ranking
    sorted_unique = sorted(list(set(averages)), reverse=True)
    pos_map = {avg: humanize.ordinal(i + 1) for i, avg in enumerate(sorted_unique)}

    # Find THIS student's exact average for ranking
    exact_avg = None
    for sid, rec in student_totals.items():
        if rec["student"].id == student.id:
            exact_avg = rec["average"]
            break

    class_position = pos_map.get(exact_avg, "-")

    # ---------------------------------------------------------
    # Fetch behaviour, signatures, header image
    # ---------------------------------------------------------
    bhv = StudentBehaviouralAssessment.objects.filter(
        session=session,
        term=term,
        student=student
    ).first()

    signs = signature.objects.all().first()
    some_images = sets.objects.first()

    # ---------------------------------------------------------
    # STRUCTURE DATA FOR TEMPLATE (same as multiple template)
    # ---------------------------------------------------------
    result_data = {
        "student": student,
        "class": clas.class_name,
        "subjects": subjects,
        "total": total_score,
        "average": average_display,  # rounded for display
        "grade": overall_grade,
        "comment": comment,
        "position": class_position,
        "bhv": bhv,
    }

    # ---------------------------------------------------------
    # RENDER SINGLE RESULT USING SAME TEMPLATE STRUCTURE
    # ---------------------------------------------------------
    return render(request, "src/print_results.html", {

        "results": [result_data],   # important: template loops over results

        # top-level info
        "session": session.session_name,
        "term": term.term_name,
        "class_name": clas.class_name,
        "no_in_class": len(student_totals),

        # design assets
        "header_image": some_images.h_image.url if some_images and some_images.h_image else None,
        "signs": signs,

        # class type flags
        "use_headteacher_sign": USE_HEADTEACHER,
        "is_kgn_nur": IS_KGN_NUR,
        "is_basic": IS_BASIC,

        "date": datetime.datetime.now().strftime("%d-%m-%Y"),
    })



def score_sheet(request):
    if request.method == 'POST' and "form1" in request.POST:
        classs = request.POST['class']
        class_id = StudentClass.objects.get(class_name=classs)
        class_id = class_id.id


        dukka = Student.objects.filter(student_class=class_id).all()
        context = {
            'classs':classs,
            'all': dukka
        }

        return render(request, 'src/score_sheet1.html', context)
    else:
        classes = StudentClass.objects.all()
        context = {
            'classes': classes
        }
        return render(request, 'src/score_sheet.html', context)





def excel_score_sheet(request):
    if request.user.is_authenticated:
        if request.method == 'POST' and "form1" in request.POST:

            classs = request.POST['class']
            subject = request.POST['subject']
            template_name = f"{classs}-{subject}"

            # Get class ID
            class_id = StudentClass.objects.get(class_name=classs).id

            # Get students in the class
            dukka = Student.objects.filter(student_class=class_id)

            # Create Excel workbook
            book = Workbook()
            sheet = book.active
            sheet.title = 'result'

            # Header row
            sheet['A1'] = "ID NO"
            sheet['B1'] = "Names"
            sheet['C1'] = "1st C.A"
            sheet['D1'] = "2nd C.A"
            sheet['E1'] = "Exams"

            y = 2

            # Populate rows
            for student in dukka:
                sheet.cell(row=y, column=1, value=student.id)
                sheet.cell(row=y, column=2, value=student.student_name)
                sheet.cell(row=y, column=3, value=0)  # CA1
                sheet.cell(row=y, column=4, value=0)  # CA2
                sheet.cell(row=y, column=5, value=0)  # Exams
                y += 1

            # Save workbook to memory using BytesIO
            output = BytesIO()
            book.save(output)
            output.seek(0)

            # Prepare response
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{template_name}.xlsx"'

            return response

        else:
            # GET request — show form
            classes = StudentClass.objects.all()
            subjects = Subject.objects.all()
            context = {
                'classes': classes,
                'subjects': subjects
            }
            return render(request, 'src/score_sheet.html', context)

    else:
        return render(request, 'src/welcome.html')


def logout(request):
    auth.logout(request)
    return redirect('/')

def not_uploaded_results(request):
    if request.user.is_authenticated:
        classes = StudentClass.objects.all()
        subjects = Subject.objects.all()
        context = {
            'subjects': subjects,
            'classes': classes
        }
        if "GET" == request.method:
            return render(request, 'src/stdupload.html', context)
        else:

            session = request.POST['session']
            term = request.POST['term']
            classs = request.POST['class']
            subject = request.POST['subject']

            #---------------Getting IDs of the Subjects, Session, and Class---------------------
            ses = Session.objects.get(session_name=session)
            session_id = ses.id

            trm = Term.objects.get(term_name=term)
            term_id = trm.id

            clas = StudentClass.objects.get(class_name=classs)
            class_id = clas.id

            subj = Subject.objects.get(subject_name=subject)
            subject_id = subj.id

            excel_file = request.FILES["excel_file"]

            # you may put validations here to check extension or file size

            wb = openpyxl.load_workbook(excel_file)

            # getting a particular sheet by name out of many sheets
            worksheet = wb["result"]
            ids = []
            ca1s = []
            ca2s = []
            exams = []
            ses_ids = []
            trm_ids = []
            clas_ids = []
            subj_ids =[]
            total_score = []
            for i, j, k, l in zip((worksheet['A']), (worksheet['C']), (worksheet['D']), (worksheet['E'])):
                ids.append(i.value)
                ca1s.append(j.value)
                ca2s.append(k.value)
                exams.append(l.value)
                ses_ids.append(session_id)
                trm_ids.append(term_id)
                clas_ids.append(class_id)
                subj_ids.append(subject_id)


            #------------removing first row of the excel-----------
            f_ids = ids[1:]
            f_ca1s = ca1s[1:]
            f_ca2s = ca2s[1:]
            f_exams = exams[1:]

            #-----------coverting to interger list----------------
            a_ids = [int(i) for i in f_ids]
            a_ca1s = [int(i) for i in f_ca1s]
            a_ca2s = [int(i) for i in f_ca2s]
            a_exams = [int(i) for i in f_exams]

            #--------------sumation of the CA to get Total Marks
            for i in range(len(f_ca1s)):
                total_score.append(a_ca1s[i] + a_ca2s[i] + a_exams[i])
            print(total_score)

            clso = int(class_id)
            sess = int(session_id)
            trm = int(term_id)
            sbj = int(subject_id)

            c = connection.cursor()





            c.executemany('INSERT INTO src_studentresult (ca1, ca2, exams, total, student_id, student_class_id, subject_id, session_id, term_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)', zip(a_ca1s, a_ca2s, a_exams, total_score, a_ids, clas_ids, subj_ids, ses_ids, trm_ids))
            connection.commit()
            c.close()
            messages.info(request, 'Result Uploaded Successfully')

            s = "ZMS/SS/12"



            update_position_query = StudentResult.objects.filter(session=session_id, term=term_id, student_class=class_id, subject=subject_id).all()

            f_total = [x.total for x in update_position_query]

            position = []
            for total in f_total:
                count = 0
                for pos in f_total:
                    if pos > total:
                        count += 1
                position.append(count + 1)

            for i, j in zip(position, update_position_query):
                j.subject_position = i
                j.save()





        return render(request,'src/stdupload.html', context)
    else:
        return render(request, 'src/welcome.html')


def bhvxl(request):
    global students
    global subject
    global term
    global session
    global classs
    global class_id
    global session_id
    global term_id
    global subject_id
    global student_id
    if request.user.is_authenticated:
        classes = StudentClass.objects.all()
        context = {
            'classes': classes
        }
        if "GET" == request.method:
            return render(request, 'src/behav.html', context)
        else:

            session = request.POST['session']
            term = request.POST['term']
            classs = request.POST['class']


            #---------------Getting IDs of the Subjects, Session, and Class---------------------
            ses = Session.objects.get(session_name=session)
            session_id = ses.id

            trm = Term.objects.get(term_name=term)
            term_id = trm.id

            clas = StudentClass.objects.get(class_name=classs)
            class_id = clas.id

            excel_file = request.FILES["excel_file"]

            # you may put validations here to check extension or file size

            wb = openpyxl.load_workbook(excel_file)

            # getting a particular sheet by name out of many sheets
            worksheet = wb["result"]
            ids = []
            ses_ids = []
            trm_ids = []
            clas_ids = []
            conduct = []
            punc = []
            ded = []
            part = []
            hosp = []
            creat = []
            phy = []
            neat = []
            school_opened = []
            days_present = []
            days_absent = []
            next_date_of_resumption = []


            for i, j, k, l, m, n, o, p, q, r, s, t, u in zip((worksheet['A']), (worksheet['C']), (worksheet['D']),
                                                          (worksheet['E']), (worksheet['F']), (worksheet['G']),
                                                          (worksheet['H']), (worksheet['I']), (worksheet['J']),
                                                          (worksheet['K']), (worksheet['L']), (worksheet['M']), (worksheet['N']),):
                ids.append(i.value)
                ses_ids.append(session_id)
                trm_ids.append(term_id)
                clas_ids.append(class_id)
                conduct.append(j.value)
                punc.append(k.value)
                ded.append(l.value)
                part.append(m.value)
                hosp.append(n.value)
                creat.append(o.value)
                phy.append(p.value)
                neat.append(q.value)
                school_opened.append(r.value)
                days_present.append(s.value)
                days_absent.append(t.value)
                next_date_of_resumption.append(u.value)




            #------------removing first row of the excel-----------
            f_ids = ids[1:]
            conduct = conduct[1:]
            punc = punc[1:]
            ded = ded[1:]
            part = part[1:]
            hosp = hosp[1:]
            creat = creat[1:]
            phy = phy[1:]
            neat = neat[1:]
            school_opened = school_opened[1:]
            days_present = days_present[1:]
            days_absent = days_absent[1:]
            next_date_of_resumption = next_date_of_resumption[1:]

            #-----------coverting to interger list----------------
            a_ids = [int(i) for i in f_ids]


            #--------------sumation of the CA to get Total Marks


            clso = int(class_id)
            sess = int(session_id)
            trm = int(term_id)


            c = connection.cursor()




            queryset = StudentBehaviouralAssessment.objects.filter(session=session_id, term=term_id, student_class=class_id).exists()
            if queryset:
                print("record is already there")
            else:
                c = connection.cursor()
                c.executemany("""INSERT INTO src_studentbehaviouralassessment (conduct, punctuality, dedication, participation, hospitality,
                    neatness, creativity, physical, school_opened, days_present, days_absent, next_date_of_resumption, session_id, student_id, student_class_id, term_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    zip(conduct, punc, ded, part, hosp, neat, creat, phy, school_opened, days_present, days_absent, next_date_of_resumption, ses_ids, f_ids, clas_ids, trm_ids))
                connection.commit()
                c.close()

            s = "ZMS/SS/12"



        return render(request,'src/behav.html', context)
    else:
        return render(request, 'src/welcome.html')
'''
    global students
    global subject
    global term
    global session
    global classs
    global class_id
    global session_id
    global term_id
    global subject_id
    global student_id
    if request.user.is_authenticated:


        if request.method == 'POST' and "form1" in request.POST:

            session = request.POST['session']
            term = request.POST['term']
            classs = request.POST['class']


            ses = Session.objects.get(session_name=session)
            session_id = ses.id

            trm = Term.objects.get(term_name=term)
            term_id = trm.id

            clas = StudentClass.objects.get(class_name=classs)
            class_id = clas.id

            students = Student.objects.filter(student_class=class_id).all()
            student_id = []

            for i in students:
                student_id.append(i.id)
            print(student_id)

            classes = StudentClass.objects.all()

            context = {
                'classes': classes,
                'students': students
                }

            return render(request, 'src/settings.html', context)

        if request.method == 'POST' and "form2" in request.POST:

            ses_ids = []
            trm_ids = []
            clas_ids = []

            ids = request.POST.getlist('id')
            conduct = request.POST.getlist('conduct')
            punc = request.POST.getlist('punctuality')
            ded = request.POST.getlist('dedication')
            part = request.POST.getlist('participation')
            hosp = request.POST.getlist('hospitality')
            creat= request.POST.getlist('creativity')
            phy = request.POST.getlist('physical')
            neat = request.POST.getlist('neatness')

            for i in student_id:
                ses_ids.append(session_id)
                trm_ids.append(term_id)
                clas_ids.append(class_id)



            queryset = StudentBehaviouralAssessment.objects.filter(session=session_id, term=term_id, student_class=class_id).exists()
            if queryset:
                print("record is already there")
            else:
                c = connection.cursor()
                c.executemany("""INSERT INTO src_studentbehaviouralassessment (conduct, punctuality, dedication, participation, hospitality,
                    neatness, creativity, physical, session_id, student_id, student_class_id, term_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    zip(conduct, punc, ded, part, hosp, neat, creat, phy, ses_ids, student_id, clas_ids, trm_ids))
                connection.commit()
                c.close()


            classes = StudentClass.objects.all()

            context = {
                'classes': classes
                }

            return render(request, 'src/settings.html', context)



        else:
            classes = StudentClass.objects.all()

            context = {
                'classes': classes
                }

            return render(request, 'src/settings.html', context)
    else:
        return render(request, 'src/welcome.html')
'''



def behav_sheet(request):
    if request.user.is_authenticated:

        if request.method == 'POST' and "form1" in request.POST:

            classs = request.POST['class']
            template_name = f"{classs} Behavioral Sheet"

            class_id = StudentClass.objects.get(class_name=classs).id
            dukka = Student.objects.filter(student_class=class_id)

            # Create workbook
            book = Workbook()
            sheet = book.active
            sheet.title = 'result'

            # Header
            sheet['A1'] = "ID NO"
            sheet['B1'] = "Names"
            sheet['C1'] = "conduct"
            sheet['D1'] = "punc."
            sheet['E1'] = "dedic."
            sheet['F1'] = "part."
            sheet['G1'] = "hosp."
            sheet['H1'] = "creat."
            sheet['I1'] = "phy."
            sheet['J1'] = "neat."
            sheet['K1'] = "school_opened"
            sheet['L1'] = "days_present"
            sheet['M1'] = "days_absent"
            sheet['N1'] = "next_date_of_resumption"

            y = 2

            # Populate rows
            for st in dukka:
                sheet.cell(row=y, column=1, value=st.id)
                sheet.cell(row=y, column=2, value=st.student_name)
                sheet.cell(row=y, column=3, value=0)
                sheet.cell(row=y, column=4, value=0)
                sheet.cell(row=y, column=5, value=0)
                sheet.cell(row=y, column=6, value=0)
                sheet.cell(row=y, column=7, value=0)
                sheet.cell(row=y, column=8, value=0)
                sheet.cell(row=y, column=9, value=0)
                sheet.cell(row=y, column=10, value=0)
                sheet.cell(row=y, column=11, value=0)
                sheet.cell(row=y, column=12, value=0)
                sheet.cell(row=y, column=13, value=0)
                sheet.cell(row=y, column=14, value=0)
                y += 1

            # Save to BytesIO instead of save_virtual_workbook
            output = BytesIO()
            book.save(output)
            output.seek(0)

            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{template_name}.xlsx"'

            return response

        else:
            classes = StudentClass.objects.all()
            return render(request, 'src/behavsheet.html', {'classes': classes})

    else:
        return render(request, 'src/welcome.html')


def std_alts(request):

    if request.user.is_authenticated:
        global f_s_id
        global context
        global student_name
        global student_class
        if request.method == 'POST' and "form1" in request.POST:
            student_id = request.POST['ID']
            f_s_id = student_id

            student = Student.objects.get(id=f_s_id)
            classes = StudentClass.objects.all()

            context = {
                "classes": classes,
                "student": student,
            }
            return render(request, "src/std-alts.html", context)

            #Adding New Students
        if request.method == 'POST' and "add" in request.POST:
            name = request.POST['name']
            classs = request.POST['class']

            Student.objects.create(student_name=name, student_class=StudentClass.objects.get(class_name=classs))
            messages.success(request, 'Record Created Successfully')
            return render(request, "src/std-alts.html", context)


            # Updating ---------------------------------
        if request.method == 'POST' and "update" in request.POST:
            student_name = request.POST['name']
            student_class = request.POST['class']
            student_status = request.POST['status']

            temp = Student.objects.get(id=f_s_id)
            temp.student_name = student_name
            temp.student_class = StudentClass.objects.get(class_name=student_class)
            #temp.status = student_status
            temp.save()
            messages.success(request, 'Record Updated Successfully')
            return render(request, "src/std-alts.html", context)

            # Deleting ------------------------
        if request.method == 'POST' and "delete" in request.POST:
            Student.objects.get(id=f_s_id).delete()
            messages.success(request, 'Record Deleted Successfully')
            return render(request, "src/std-alts.html", context)


        if request.method == 'POST' and "save" in request.POST:
            current_class = request.POST['current_class']
            promotion_class = request.POST['promotion_class']

            #-----getting classes ids---------
            c_clas = StudentClass.objects.get(class_name=current_class)
            c_class_id = c_clas.id

            p_clas = StudentClass.objects.get(class_name=promotion_class)
            #p_class_id = p_clas.id

            #-----------getting student in a class---------------
            students_in_class = Student.objects.filter(student_class=c_class_id).all()
            session_get = Session.objects.all()
            real_session = session_get[(len(session_get)-2)]
            for student in students_in_class:
                if 'SS 3' not in c_clas.class_name:
                    student_results = StudentResult.objects.filter(student=student.id, term=3, session=real_session.id).all()
                    total = []
                    for result in student_results:
                        total.append(result.total)

                    avr = sum(total)/len(total)
                    if avr >= 40:


                        query = student
                        query.student_class = p_clas
                        query.save()
                else:
                    student.student_class = p_clas
                    student.save()
            classes = StudentClass.objects.all()
            context = {
                "classes": classes,
            }

            return render(request, "src/std-alts.html", context)


        else:
            classes = StudentClass.objects.all()
            context = {
                "classes": classes,
            }
            return render(request, "src/std-alts.html", context)


    else:
        return render(request, 'src/welcome.html')

def class_summary(request):
    if request.user.is_authenticated:
        if request.method == 'POST' and "form1" in request.POST:
            classs = request.POST['class']
            term = request.POST['term']
            session = request.POST['session']
            subject = request.POST['subject']

            result_summary = StudentResult.objects.filter(subject=Subject.objects.get(subject_name=subject), student_class=StudentClass.objects.get(class_name=classs),
                            session=Session.objects.get(session_name=session), term=Term.objects.get(term_name=term)).all()


            total = []
            for i in result_summary:
                total.append(i.total)


            a_count = 0
            b_count = 0
            c_count = 0
            d_count = 0
            e_count = 0
            f_count = 0
            total_students = len(total)

            for i in total:
                if i >= 70:
                    a_count = a_count + 1
                elif i >= 60:
                    b_count = b_count + 1
                elif i >= 50:
                    c_count = c_count + 1
                elif i >=45:
                    d_count = d_count + 1
                elif i >=40:
                    e_count = e_count + 1
                else:
                    f_count = f_count + 1
            print(f_count)

            context = {
                'a_count': a_count,
                'b_count': b_count,
                'c_count': c_count,
                'd_count': d_count,
                'e_count': e_count,
                'f_count': f_count,
                'summary': result_summary,
                'total_students': total_students,
                'subject': subject,
                'classs': classs
            }




            '''total = []
            subject = []
            for result in result_summary:
                subject.append(result.subject)
                total.append(result.total)

            final_subjects = list(dict.fromkeys(subject))
            length_resulut_summary = len(result_summary)
            length_final_subjects = len(final_subjects)
            number_of_students = length_resulut_summary/length_final_subjects
            splits = np.array_split(total, length_final_subjects)

            for array in splits:
                print(array)'''




            return render(request, 'src/summary_page.html', context)
        else:
            classes = StudentClass.objects.all()
            terms = Term.objects.all()
            sessions = Session.objects.all()
            subjects = Subject.objects.all()
            context = {
                'classes': classes,
                'terms': terms,
                'sessions': sessions,
                'subjects': subjects
            }
            return render(request, 'src/class_summary.html', context)
        pass
    else:
        return render(request, 'src/welcome.html')

def delete_result(request):
    if request.user.is_authenticated:
        if request.method == 'POST' and "form1" in request.POST:
            session = request.POST['session']
            term = request.POST['term']
            classs = request.POST['class']

            ses = Session.objects.get(session_name=session)
            session_id = ses.id

            trm = Term.objects.get(term_name=term)
            term_id = trm.id

            clas = StudentClass.objects.get(class_name=classs)
            class_id = clas.id

            # subj = Subject.objects.get(subject_name=subject)
            # subject_id = subj.id

            query = StudentResult.objects.filter(session=session_id, term=term_id, student_class=class_id).all()
            query.delete()

            messages.info(request, 'Result Deleted Successfully')

            subjects = Subject.objects.all()
            classes = StudentClass.objects.all()
            sessions = Session.objects.all()
            terms = Term.objects.all()
            context = {
                'subjects': subjects,
                'classes': classes,
                'sessions': sessions,
                'terms': terms,
            }

            return render(request, 'src/resultdelete.html', context)

        if request.method == 'POST' and "form2" in request.POST:
            session = request.POST['session']
            term = request.POST['term']
            classs = request.POST['class']
            subject = request.POST['subject']

            ses = Session.objects.get(session_name=session)
            session_id = ses.id

            trm = Term.objects.get(term_name=term)
            term_id = trm.id

            clas = StudentClass.objects.get(class_name=classs)
            class_id = clas.id

            subj = Subject.objects.get(subject_name=subject)
            subject_id = subj.id

            query = StudentResult.objects.filter(session=session_id, term=term_id, student_class=class_id, subject=subject_id).all()
            query.delete()

            messages.info(request, 'Result Deleted Successfully')

            subjects = Subject.objects.all()
            classes = StudentClass.objects.all()
            sessions = Session.objects.all()
            terms = Term.objects.all()
            context = {
                'subjects': subjects,
                'classes': classes,
                'sessions': sessions,
                'terms': terms,
            }

            return render(request, 'src/resultdelete.html', context)

        else:
            subjects = Subject.objects.all()
            classes = StudentClass.objects.all()
            sessions = Session.objects.all()
            terms = Term.objects.all()
            context = {
                'subjects': subjects,
                'classes': classes,

            }
            return render(request, "src/resultdelete.html", context)
    else:
        return render(request, 'src/welcome.html')


def delete_bhv(request):
    if request.user.is_authenticated:
        if request.method == 'POST' and "form1" in request.POST:
            session = request.POST['session']
            term = request.POST['term']
            classs = request.POST['class']

            ses = Session.objects.get(session_name=session)
            session_id = ses.id

            trm = Term.objects.get(term_name=term)
            term_id = trm.id

            clas = StudentClass.objects.get(class_name=classs)
            class_id = clas.id

            # subj = Subject.objects.get(subject_name=subject)
            # subject_id = subj.id

            query = StudentBehaviouralAssessment.objects.filter(session=session_id, term=term_id, student_class=class_id).all()
            query.delete()

            messages.info(request, 'Result Deleted Successfully')


            classes = StudentClass.objects.all()
            sessions = Session.objects.all()
            terms = Term.objects.all()
            context = {

                'classes': classes,
                'sessions': sessions,
                'terms': terms,
            }

            return render(request, 'src/bhvdelete.html', context)


        else:

            classes = StudentClass.objects.all()
            sessions = Session.objects.all()
            terms = Term.objects.all()
            context = {

                'classes': classes,

            }
            return render(request, "src/bhvdelete.html", context)
    else:
        return render(request, 'src/welcome.html')

def bhvxl_missed(request):
    global students
    global subject
    global term
    global session
    global classs
    global class_id
    global session_id
    global term_id
    global subject_id
    global student_id
    if request.user.is_authenticated:
        classes = StudentClass.objects.all()
        context = {
            'classes': classes
        }
        if "GET" == request.method:
            return render(request, 'src/behav.html', context)
        else:

            session = request.POST['session']
            term = request.POST['term']
            classs = request.POST['class']


            #---------------Getting IDs of the Subjects, Session, and Class---------------------
            ses = Session.objects.get(session_name=session)
            session_id = ses.id

            trm = Term.objects.get(term_name=term)
            term_id = trm.id

            clas = StudentClass.objects.get(class_name=classs)
            class_id = clas.id

            excel_file = request.FILES["excel_file"]

            # you may put validations here to check extension or file size

            wb = openpyxl.load_workbook(excel_file)

            # getting a particular sheet by name out of many sheets
            worksheet = wb["result"]
            ids = []
            ses_ids = []
            trm_ids = []
            clas_ids = []
            conduct = []
            punc = []
            ded = []
            part = []
            hosp = []
            creat = []
            phy = []
            neat = []
            school_opened = []
            days_present = []
            days_absent = []


            for i, j, k, l, m, n, o, p, q, r, s, t in zip((worksheet['A']), (worksheet['C']), (worksheet['D']),
                                                          (worksheet['E']), (worksheet['F']), (worksheet['G']),
                                                          (worksheet['H']), (worksheet['I']), (worksheet['J']),
                                                          (worksheet['K']), (worksheet['L']), (worksheet['M']),):
                ids.append(i.value)
                ses_ids.append(session_id)
                trm_ids.append(term_id)
                clas_ids.append(class_id)
                conduct.append(j.value)
                punc.append(k.value)
                ded.append(l.value)
                part.append(m.value)
                hosp.append(n.value)
                creat.append(o.value)
                phy.append(p.value)
                neat.append(q.value)
                school_opened.append(r.value)
                days_present.append(s.value)
                days_absent.append(t.value)




            #------------removing first row of the excel-----------
            f_ids = ids[1:]
            conduct = conduct[1:]
            punc = punc[1:]
            ded = ded[1:]
            part = part[1:]
            hosp = hosp[1:]
            creat = creat[1:]
            phy = phy[1:]
            neat = neat[1:]
            school_opened = school_opened[1:]
            days_present = days_present[1:]
            days_absent = days_absent[1:]

            #-----------coverting to interger list----------------
            a_ids = [int(i) for i in f_ids]


            #--------------sumation of the CA to get Total Marks


            clso = int(class_id)
            sess = int(session_id)
            trm = int(term_id)


            c = connection.cursor()





            c.executemany("""INSERT INTO src_studentbehaviouralassessment (conduct, punctuality, dedication, participation, hospitality,
                neatness, creativity, physical, school_opened, days_present, days_absent, session_id, student_id, student_class_id, term_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                zip(conduct, punc, ded, part, hosp, neat, creat, phy, school_opened, days_present, days_absent, ses_ids, f_ids, clas_ids, trm_ids))
            connection.commit()
            c.close()

            s = "ZMS/SS/12"



        return render(request,'src/behav.html', context)
    else:
        return render(request, 'src/welcome.html')

def class_result_summary(request):
    global class_id
    global session_id
    global term_id

    if request.user.is_authenticated:

        if request.method == 'POST' and "form1" in request.POST:
            classes = StudentClass.objects.all()

            context = {
                'classes': classes
                }

            session = request.POST['session']
            term = request.POST['term']
            classs = request.POST['class']

            directorAcadamic = "Director Acadamic & Examination"
            if "SS" in classs or "JS" in classs:
                pcHeadTeacher = "Principal"

            else:
                pcHeadTeacher = "Head Teacher"

            ses = Session.objects.get(session_name=session)
            session_id = ses.id

            trm = Term.objects.get(term_name=term)
            term_id = trm.id

            clas = StudentClass.objects.get(class_name=classs)
            class_id = clas.id

            results = StudentResult.objects.filter(session=session_id, term=term_id, student_class=class_id)
            bulk = {}


            eachSubject = []
            for result in results:
                ca1_total = 0
                ca2_total = 0
                exams_total = 0
                total = 0
                subjects = []
                for subject in results:
                    if subject.student == result.student:
                        subjects.append(subject)
                        eachSubject.append(subject.subject)
                        ca1_total += subject.ca1
                        ca2_total += subject.ca2
                        exams_total += subject.exams
                        total += subject.total
                        subject_pos = subject.subject_position
                        grade = subject.grade
                        terms = subject.term
                        sesss = subject.session
                        classs = subject.student_class
                        avr = total/len(subjects)



                    bulk[result.student] = {
                    "student": result.student,
                    "subjects": subjects,
                    "ca1_total": ca1_total,
                    "ca2_total": ca1_total,
                    "exams_total": exams_total,
                    "total": total,
                    "grade": grade,
                    "subject_pos": subject_pos,
                    "terms": terms,
                    "sesss": sesss,
                    "classs": classs,
                    "overal_total": total,
                    "average": round(avr, 2),




                    }

            classes = StudentClass.objects.all()
            bhv = StudentBehaviouralAssessment.objects.filter(session=session_id, term=term_id, student_class=class_id)
            signs = signature.objects.filter(classs=classs).first()
            some_images = sets.objects.first()

            f_classs = str(classs)
            date = datetime.datetime.now()
            eachSubject = list(dict.fromkeys(eachSubject))
            context = {
                'date': date,
                'f_classs': f_classs,
                'classes': classes,
                'results': bulk,
                'bhv': bhv,
                'signs': signs,
                'some_images':some_images,
                'p_term': term,
                'eachSubject': eachSubject,
                "session": session,
                "term": term,
                "class": classs,
                "directorAcadamic": directorAcadamic,
                "pcHeadTeacher": pcHeadTeacher,
            }



            return render(request, 'src/class_result_summary.html', context)

        else:
            classes = StudentClass.objects.all()

            context = {
                'classes': classes
                }

            return render(request, 'src/result_view.html', context)

    else:
        return render(request, 'src/welcome.html')